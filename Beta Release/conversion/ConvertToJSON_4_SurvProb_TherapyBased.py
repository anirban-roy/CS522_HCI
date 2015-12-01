__author__ = 'Anirban'
#This generates the JSON for drill down treemap visualization against the demographics data
# Here the data is generated such that the average survival probability is used as the treemap display parameter.
import openpyxl
wb = openpyxl.load_workbook('SurvivalProbability.xlsx')
ethnicity = dict()
ethnicityCount = dict()
site = dict()
sex = dict()
sheet = wb.get_sheet_by_name('SurvivalProbability')
max_row = sheet.max_row
max_col = sheet.max_column
print max_row, max_col
for i in range(2, max_row,1):
    #for j in range(1,max_col,1):
    #[item] = frequencies.get(item, 0) + 1
    c_site = str(sheet.cell(row=i, column=3).value)
    c_ethni = str(sheet.cell(row=i, column=2).value)
    c_sex = str(sheet.cell(row=i, column=4).value)
    c_probability = float(sheet.cell(row=i, column=9).value)
    c_age = float(sheet.cell(row=i, column=1).value)
    c_therapy = str(sheet.cell(row=i, column=7).value)

    if ethnicity.get(c_ethni) != None: #if the ethnicity is not new from dataset
        # therapy = ethnicity.get(c_ethni)[0]
        # if therapy.get(c_therapy, None) == None:
        #     sex = dict()
        #     sex[c_sex] = (1, c_probability, c_age)
        #     site[c_site] = (sex, 1, c_probability, c_age)
        #     therapy[c_therapy] = (site,1,c_probability, c_age)

        site = ethnicity.get(c_ethni)[0]
        #start Here #####
        if site.get(c_site, None) == None: # if the site is a new site from dataset
            therapy = dict()
            therapy[c_therapy] = (1,c_probability, c_age)
            sex = dict()
            sex[c_sex] = (therapy, 1, c_probability, c_age)
            site[c_site] = (sex, 1, c_probability, c_age)
        else:
            sex = site.get(c_site, None)[0]
            if sex.get(c_sex, None) == None:
                therapy = dict()
                therapy[c_therapy] = (1,c_probability, c_age)
                sex[c_sex] = (therapy, 1, c_probability, c_age)
            else:
                therapy = sex.get(c_sex, None)[0]
                if therapy.get(c_therapy, None) == None:
                    therapy[c_therapy] = (1, c_probability, c_age)
                else:
                    therapy[c_therapy] = (therapy.get(c_therapy)[0] + 1,
                                          therapy.get(c_therapy)[1] + c_probability,
                                          therapy.get(c_therapy)[2] + c_age)

                sex[c_sex] = (therapy, sex.get(c_sex)[1]+1,
                                 sex.get(c_sex)[2] + c_probability,
                                 sex.get(c_sex)[3] + c_age)
            site[c_site] = (sex, site.get(c_site)[1]+1,
                            site.get(c_site)[2] + c_probability,
                            site.get(c_site)[3] + c_age)

        ethnicity[c_ethni] = (site, ethnicity.get(c_ethni)[1] + 1,
                              ethnicity.get(c_ethni)[2] + c_probability,
                              ethnicity.get(c_ethni)[3] + c_age)
        #ethnicityCount[c_ethni] = ethnicityCount.get(c_ethni,0) + 1
    else:
        therapy = dict()
        therapy[c_therapy] = (1,c_probability, c_age)
        sex = dict()
        sex[c_sex] = (therapy, 1, c_probability, c_age)
        site = dict()
        site[c_site] = (sex, 1, c_probability, c_age)
        ethnicity[c_ethni] = (site, 1, c_probability, c_age)
        ethnicityCount[c_ethni] = 1


outputEth = open('demographics_ethnicity_survivalprobab.json', 'w')
outputEth.write('{\"data\":[')
i=0
outputString = ""
eth_length = len(ethnicity)
eth_index = 1
for ethni in ethnicity:
    count = ethnicity.get(ethni)[1]
    avgSurvProbab = (ethnicity.get(ethni)[2]/count) * 100
    avgAge = ethnicity.get(ethni)[3]/count
    outputString += "{\"ethnicity\":\"" + ethni + "\", \"count\": \"" + str(count) + "\", \"avgsurvprob\": \"" + str('%.2f' % avgSurvProbab)+ "\", \"avgage\":\"" + str('%.2f' % avgAge) + "\", \"sites\":["
    sites = ethnicity.get(ethni)[0]
    si_length = len(sites)
    si_index = 1
    for site in sites:
        count = sites.get(site)[1]
        avgSurvProbab = (sites.get(site)[2]/count) * 100
        avgAge = sites.get(site)[3]/count
        outputString += "{\"site\":\"" + site + "\", \"count\": \"" + str(count) + "\", \"avgsurvprob\": \"" + str('%.2f' % avgSurvProbab)+ "\", \"avgage\":\"" + str('%.2f' % avgAge) + "\", \"sexes\":["
        sexes = sites.get(site)[0]
        se_length = len(sexes)
        se_index = 1
        for sex in sexes:
            count = sexes.get(sex)[1]
            avgSurvProbab = (sexes.get(sex)[2]/count) * 100
            avgAge = sexes.get(sex)[3]/count
            outputString += "{\"sex\":\"" + sex + "\", \"count\": \"" + str(count) + "\", \"avgsurvprob\": \"" + str('%.2f' % avgSurvProbab)+ "\", \"avgage\":\"" + str('%.2f' % avgAge) + "\", \"therapies\":["
            therapies = sexes.get(sex)[0]
            th_length = len(therapies)
            th_index = 1
            for therapy in therapies:
                count = therapies.get(therapy)[0]
                avgSurvProbab = (therapies.get(therapy)[1]/count) * 100
                avgAge = therapies.get(therapy)[2]/count
                outputString += "{\"name\":\"" + therapy + "\", \"count\": \"" + str(count) + "\", \"value\": " + str('%.2f' % avgSurvProbab)+ ", \"avgage\":\"" + str('%.2f' % avgAge) + "\"}" # , \"color\": \"color(\'"+therapy+"\')\"
                if th_index < th_length: # append a , if this is not last entry in therapies dictionary
                    outputString += ","
                th_index += 1

            outputString += "]}" #end of therapy block array
            if se_index < se_length: # append a , if this is not last entry in sexes dictionary
                outputString += ","
            se_index += 1

        outputString += "]}" #end of sites block array
        if si_index < si_length: # append a , if this is not last entry in sites dictionary
            outputString += ","
        si_index += 1
    outputString += "]}" #end of sites block array
    if eth_index < eth_length: # append a , if this is not last entry in sites dictionary
        outputString += ","
    eth_index += 1

outputString += "]}" #end of ethnicity block array
#outputString += "]}" #end of data block array

outputEth.write(outputString)
outputEth.close()


# The below code used for generating JSON string to handle drill down events in the Tree Map
# for ethni in ethnicity:
#     count = ethnicity.get(ethni)[1]
#     avgAge = ethnicity.get(ethni)[3]/count
#     avgSurvProbab = (ethnicity.get(ethni)[2]/count) * 100
#     outputEth.write("{\"id\":\""+ethni+"\",\"name\":\""+ethni+"\", \"color\": color(\""+ethni+"\"), \"value\":" + str('%.2f' % avgSurvProbab)+ ", \"avgage\": \""+ str('%.2f' % avgAge) +"\", \"drilldown\":\""+ethni+"\"},\n")
#
# outputEth.write(']}')
# outputEth.close()
# outputDrill = open('demographics_survivalprobab_drilldown.json', 'w')
# outputDrill.write('{\"data\":[')
# for ethni in ethnicity:
#     sites = ethnicity.get(ethni)[0]
#     for site in sites:
#         count = sites.get(site)[1]
#         avgAge = sites.get(site)[3]/count
#         avgSurvProbab = (sites.get(site)[2]/count) * 100
#         outputDrill.write("{\"id\":\""+ethni+site+"\", \"name\":\""+site+"\", \"value\":" +str('%.2f' % avgSurvProbab) +", \"avgage\": "+ str('%.2f' % avgAge) + "},\n")
#         sexes = sites.get(site)[0]
#         for sex in sexes:
#             count = sexes.get(sex)[0]
#             avgAge = sexes.get(sex)[2]/count
#             avgSurvProbab = (sexes.get(sex)[1]/count) * 100
#             if sex.lower() == 'male':
#                 outputDrill.write("{\"id\":\""+ethni+site+sex+"\",\"parent\":\""+ethni+site+"\",\"name\":\""+sex+"\", \"value\":"+str('%.2f' % avgSurvProbab) +", \"avgage\": "+ str('%.2f' % avgAge)+", \"color\": color(\"male\")},\n")
#             else:
#                 outputDrill.write("{\"id\":\""+ethni+site+sex+"\",\"parent\":\""+ethni+site+"\",\"name\":\""+sex+"\", \"value\":" +str('%.2f' % avgSurvProbab) +", \"avgage\": "+ str('%.2f' % avgAge)+ ",\"color\": color(\"female\")},\n")
#     outputDrill.write('\n\n')
# outputDrill.write("]}")
# outputDrill.close()
# # print ethnicity

