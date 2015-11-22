__author__ = 'Anirban'
#This generates the JSON for drill down treemap visualization against the demographics data
import openpyxl
wb = openpyxl.load_workbook('data.xlsx')
ethnicity = dict()
ethnicityCount = dict()
site = dict()
sex = dict()
sheet = wb.get_sheet_by_name('T3T4')
max_row = sheet.max_row
max_col = sheet.max_column
print max_row, max_col
for i in range(2, max_row,1):
    #for j in range(1,max_col,1):
    #[item] = frequencies.get(item, 0) + 1
    c_site = str(sheet.cell(row=i, column=18).value)
    c_ethni = str(sheet.cell(row=i, column=28).value)
    c_sex = str(sheet.cell(row=i, column=17).value)

    if ethnicity.get(c_ethni) != None:
        site = ethnicity.get(c_ethni)
        if site == None:
            site = dict()
            ethnicity[c_ethni] = site
        sex = site.get(c_site)
        if sex == None:
            sex = dict()
            sex[c_sex] = 1
            site[c_site] = sex
        else:
            sex[c_sex] = sex.get(c_sex,0) + 1
        ethnicityCount[c_ethni] = ethnicityCount.get(c_ethni,0) + 1
    else:
        sex = dict()
        sex[c_sex] = 1
        site = dict()
        site[c_site] = sex
        #ethnicity = dict()
        ethnicity[c_ethni] = site
        ethnicityCount[c_ethni] = 1

outputEth = open('demographics_ethnicity.json', 'w')
outputEth.write('[')
for ethni in ethnicityCount:
    outputEth.write("{id:\'"+ethni+"\',"+"name:\'"+ethni+"\', color: color(\'"+ethni+"\'), value:" + str(ethnicityCount[ethni])+", drilldown:\'"+ethni+"\'},\n")

outputEth.write(']')
outputEth.close()
outputDrill = open('demographics_drilldown.json', 'w')
outputDrill.write('[')
for ethni in ethnicity:
    sites = ethnicity[ethni]
    for site in sites:
        outputDrill.write("{id:\'"+ethni+site+"\', name:\'"+site+"\'},\n")
        sexes = sites[site]
        for sex in sexes:
            if sex.lower() == 'male':
                outputDrill.write("{id:\'"+ethni+site+sex+"\',parent:\'"+ethni+site+"\',name:\'"+sex+"\', value:"+str(sexes[sex])+",color: color(\'male\')},\n")
            else:
                outputDrill.write("{id:\'"+ethni+site+sex+"\',parent:\'"+ethni+site+"\',name:\'"+sex+"\', value:"+str(sexes[sex])+",color: color(\'female\')},\n")
    outputDrill.write('\n\n')
outputDrill.write("]")
outputDrill.close()
# print ethnicity

