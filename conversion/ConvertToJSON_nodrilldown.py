__author__ = 'Anirban'
import openpyxl
wb = openpyxl.load_workbook('data.xlsx')
ethnicity = dict()
site = dict()
sex = dict()
sheet = wb.get_sheet_by_name('T3T4')
max_row = sheet.max_row
max_col = sheet.max_column
print max_row, max_col
for i in range(2, max_row,1):
    #for j in range(1,max_col,1):
    #[item] = frequencies.get(item, 0) + 1
    c_site = str(sheet.cell(row=i, column=18).value)
    c_ethni = str(sheet.cell(row=i, column=28).value)
    c_sex = str(sheet.cell(row=i, column=17).value)

    if ethnicity.get(c_ethni) != None:
        site = ethnicity.get(c_ethni)
        if site == None:
            site = dict()
            ethnicity[c_ethni] = site
        sex = site.get(c_site)
        if sex == None:
            sex = dict()
            sex[c_sex] = 1
            site[c_site] = sex
        else:
            sex[c_sex] = sex.get(c_sex,0) + 1
    else:
        sex = dict()
        sex[c_sex] = 1
        site = dict()
        site[c_site] = sex
        #ethnicity = dict()
        ethnicity[c_ethni] = site

output = open('demographics.json', 'w')
output.write('[')
for ethni in ethnicity:
    output.write("{id:\'"+ethni+"\',"+"name:\'"+ethni+"\'},")
    sites = ethnicity[ethni]
    for site in sites:
        output.write("{id:\'"+ethni+site+"\',parent:\'"+ethni+"\',name:\'"+site+"\'},")
        sexes = sites[site]
        for sex in sexes:
            if sex.lower() == 'male':
                output.write("{id:\'"+ethni+site+sex+"\',parent:\'"+ethni+site+"\',name:\'"+sex+"\', value:"+str(sexes[sex])+",color: \'#1570E8\'},")
            else:
                output.write("{id:\'"+ethni+site+sex+"\',parent:\'"+ethni+site+"\',name:\'"+sex+"\', value:"+str(sexes[sex])+",color: \'#E81515\'},")

output.write("]")
print ethnicity

