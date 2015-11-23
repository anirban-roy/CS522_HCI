__author__ = 'Anirban'
#This generates the JSON for drill down treemap visualization against the demographics data
# Here the data is generated such that the average survival probability is used as the treemap display parameter.
import openpyxl
wb = openpyxl.load_workbook('SurvivalProbability.xlsx')
ethnicity = dict()
ethnicityCount = dict()
site = dict()
sex = dict()
sheet = wb.get_sheet_by_name('SurvivalProbability')
max_row = sheet.max_row
max_col = sheet.max_column
print max_row, max_col
for i in range(2, max_row,1):
    #for j in range(1,max_col,1):
    #[item] = frequencies.get(item, 0) + 1
    c_site = str(sheet.cell(row=i, column=3).value)
    c_ethni = str(sheet.cell(row=i, column=2).value)
    c_sex = str(sheet.cell(row=i, column=4).value)
    c_probability = float(sheet.cell(row=i, column=9).value)
    c_age = float(sheet.cell(row=i, column=1).value)

    if ethnicity.get(c_ethni) != None:
        site = ethnicity.get(c_ethni)[0]
        if site.get(c_site, None) == None:
            sex = dict()
            sex[c_sex] = (1, c_probability, c_age)
            site[c_site] = (sex, 1, c_probability, c_age)
        else:
            sex = site.get(c_site,None)[0]
            if sex.get(c_sex, None) == None:
                #sex = dict()
                sex[c_sex] = (1, c_probability, c_age)
            else:
                sex[c_sex] = (sex.get(c_sex)[0] + 1, sex.get(c_sex)[1] + c_probability, sex.get(c_sex)[2] + c_age)

            site[c_site] = (sex, site.get(c_site)[1]+1,
                            site.get(c_site)[2] + c_probability,
                            site.get(c_site)[3] + c_age)

        ethnicity[c_ethni] = (site,  ethnicity.get(c_ethni)[1] + 1,
                              ethnicity.get(c_ethni)[2] + c_probability,
                              ethnicity.get(c_ethni)[3] + c_age)
        #ethnicityCount[c_ethni] = ethnicityCount.get(c_ethni,0) + 1
    else:
        sex = dict()
        sex[c_sex] = (1, c_probability, c_age)
        site = dict()
        site[c_site] = (sex, 1, c_probability, c_age)
        ethnicity[c_ethni] = (site, 1, c_probability, c_age)
        ethnicityCount[c_ethni] = 1


outputEth = open('demographics_ethnicity_survivalprobab.json', 'w')
outputEth.write('[')
for ethni in ethnicity:
    count = ethnicity.get(ethni)[1]
    avgAge = ethnicity.get(ethni)[3]/count
    avgSurvProbab = (ethnicity.get(ethni)[2]/count) * 100
    outputEth.write("{id:\'"+ethni+"\',"+"name:\'"+ethni+"\', color: color(\'"+ethni+"\'), value:" + str('%.2f' % avgSurvProbab)+ ", avgage: "+ str('%.2f' % avgAge) +", drilldown:\'"+ethni+"\'},\n")

outputEth.write(']')
outputEth.close()
outputDrill = open('demographics_survivalprobab_drilldown.json', 'w')
outputDrill.write('[')
for ethni in ethnicity:
    sites = ethnicity.get(ethni)[0]
    for site in sites:
        count = sites.get(site)[1]
        avgAge = sites.get(site)[3]/count
        avgSurvProbab = (sites.get(site)[2]/count) * 100
        outputDrill.write("{id:\'"+ethni+site+"\', name:\'"+site+"\', value:" +str('%.2f' % avgSurvProbab) +", avgage: "+ str('%.2f' % avgAge) + "},\n")
        sexes = sites.get(site)[0]
        for sex in sexes:
            count = sexes.get(sex)[0]
            avgAge = sexes.get(sex)[2]/count
            avgSurvProbab = (sexes.get(sex)[1]/count) * 100
            if sex.lower() == 'male':
                outputDrill.write("{id:\'"+ethni+site+sex+"\',parent:\'"+ethni+site+"\',name:\'"+sex+"\', value:"+str('%.2f' % avgSurvProbab) +", avgage: "+ str('%.2f' % avgAge)+", color: color(\'male\')},\n")
            else:
                outputDrill.write("{id:\'"+ethni+site+sex+"\',parent:\'"+ethni+site+"\',name:\'"+sex+"\', value:" +str('%.2f' % avgSurvProbab) +", avgage: "+ str('%.2f' % avgAge)+ ",color: color(\'female\')},\n")
    outputDrill.write('\n\n')
outputDrill.write("]")
outputDrill.close()
# # print ethnicity

